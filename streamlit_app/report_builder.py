"""Builds the Agile-style Test Execution Report for one suite.

Two independent pieces, kept in one file since they share the same data
shaping and this module has no Streamlit dependency (it runs standalone in
CI, invoked as `python report_builder.py <slug>`, as well as being imported
by the dashboard for the in-app "Test Execution Report" tab):

1. `compute_report_tables()` -- pure, deterministic derived data (defect
   summary, test execution summary, defect tracking rows, test execution
   rows) computed from a suite's already-generated
   streamlit_app/data/<slug>-test-results.json. No LLM involved.
2. `build_workbook()` -- renders that data plus a narrative dict (Test
   Objectives / Key Findings / Recommendations / Conclusion, authored by
   Claude Code as a separate pipeline step -- see
   streamlit_app/data/<slug>-test-report.json) into an .xlsx modeled on
   this project's adopted "Agile Test Report" template, using openpyxl so
   the charts are real Excel charts, not images.

Fields the pipeline doesn't actually track (defect lifecycle status,
owner, date-detected, a "Blocked" test outcome) are filled with sensible
fixed defaults rather than omitted, so the report keeps the template's
full shape -- see DEFAULT_DEFECT_STATUS / DEFAULT_OWNER below for exactly
what's a default versus real measured data.
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

REPO_ROOT = Path(__file__).parent.parent

DEFAULT_DEFECT_STATUS = "Open"  # this pipeline doesn't track defect resolution over time
DEFAULT_OWNER = "QA Automation"
DEFAULT_TESTED_BY = "Playwright Automation"

SEVERITY_TO_DISTRIBUTION = {
    "critical": "Critical",
    "high": "Major",
    "medium": "Major",
    "low": "Minor",
    "info": "Minor",
}

STEP_MARKER_RE = re.compile(r"^\s*//\s*(\d+)\.\s*(.+)$")
EXPECT_MARKER_RE = re.compile(r"^\s*//\s*(?:expect|verify)\b\s*:?\s*(.+)$", re.IGNORECASE)
PLAN_TC_HEADING_RE = re.compile(r"^####\s+.*?\b(TC-[A-Z0-9]+-\d+)\b", re.MULTILINE)
PLAN_STEP_RE = re.compile(r"^(\d+)\.\s+(.+)$")
PLAN_EXPECT_RE = re.compile(r"^\s*-\s*expect:\s*(.+)$", re.IGNORECASE)
PLAN_OBJECTIVES_RE = re.compile(
    r"###\s*1\.1\s*Test Plan Objectives\s*\n+(.+?)(?=\n##|\Z)", re.DOTALL
)


def parse_test_plan_steps(plan_text: str) -> dict[str, list[str]]:
    """Maps each TC-ID to its plain-text Steps: lines (no expect: bullets).

    A lighter-weight sibling of app.py's parse_test_plan_steps -- this one
    flattens each step to a single "N. text" string per test case (for a
    spreadsheet cell), rather than the {number, text, expectations} dicts
    the dashboard's inline view needs.
    """
    headings = list(PLAN_TC_HEADING_RE.finditer(plan_text))
    result: dict[str, list[str]] = {}
    for i, m in enumerate(headings):
        tc_id = m.group(1)
        start = m.end()
        end = headings[i + 1].start() if i + 1 < len(headings) else len(plan_text)
        block = plan_text[start:end]
        steps_idx = block.find("**Steps:**")
        if steps_idx == -1:
            continue
        lines = []
        for line in block[steps_idx:].splitlines():
            step_match = PLAN_STEP_RE.match(line)
            if step_match:
                lines.append(f"{step_match.group(1)}. {step_match.group(2).strip()}")
        if lines:
            result[tc_id] = lines
    return result


def parse_test_plan_objective(plan_text: str) -> str:
    """Pulls Section 1.1's free-text objective out of the 16-section test plan."""
    m = PLAN_OBJECTIVES_RE.search(plan_text)
    return m.group(1).strip() if m else ""


def _severity_bucket(severity: str) -> str:
    return SEVERITY_TO_DISTRIBUTION.get((severity or "").strip().lower(), "Minor")


def compute_report_tables(results_data: dict, plan_steps: dict[str, list[str]] | None = None) -> dict:
    """Pure transform: results_data (a <slug>-test-results.json's content) ->
    every deterministic table/summary the report needs. No narrative text --
    that's supplied separately (see build_report_payload)."""
    meta = results_data.get("meta", {})
    summary = results_data.get("summary", {})
    tests = results_data.get("tests", [])
    defects = results_data.get("defects", [])
    plan_steps = plan_steps or {}

    report_date = meta.get("report_date", "")

    defect_status_counts = {"Open": len(defects), "In Progress": 0, "Closed": 0}
    defect_severity_counts = {"Minor": 0, "Major": 0, "Critical": 0}
    defect_tracking_rows = []
    for defect in defects:
        bucket = _severity_bucket(defect.get("severity", ""))
        defect_severity_counts[bucket] += 1
        defect_tracking_rows.append(
            {
                "id": defect.get("id", ""),
                "date_detected": report_date,
                "description": defect.get("title", ""),
                "status": DEFAULT_DEFECT_STATUS,
                "severity": bucket,
                "owner": DEFAULT_OWNER,
                "remarks": defect.get("actual", ""),
            }
        )

    passed = int(summary.get("passed", 0))
    failed = int(summary.get("failed", 0))
    not_executed = int(summary.get("skipped", 0))
    blocked = 0  # this pipeline has no "blocked" execution outcome

    outcome_map = {"pass": "PASSED", "fail": "FAILED", "skip": "NOT EXECUTED"}
    test_execution_rows = []
    for test in tests:
        outcome = outcome_map.get(str(test.get("chromium", "")).strip().lower(), "NOT EXECUTED")
        matching_defect = next(
            (d for d in defects if test.get("id", "") in [r.strip() for r in str(d.get("test_ref", "")).split(",")]),
            None,
        )
        steps = plan_steps.get(test.get("id", ""), [])
        test_execution_rows.append(
            {
                "id": test.get("id", ""),
                "date": report_date,
                "description": test.get("title", ""),
                "steps": "\n".join(steps),
                "expected": matching_defect["expected"] if matching_defect else "All assertions pass",
                "actual": matching_defect["actual"] if matching_defect else ("All assertions passed" if outcome == "PASSED" else ""),
                "status": outcome,
                "tested_by": DEFAULT_TESTED_BY,
                "remarks": (matching_defect or {}).get("title", "") if outcome == "FAILED" else "",
            }
        )

    return {
        "meta": meta,
        "total_defects": len(defects),
        "total_tests": summary.get("test_cases", len(tests)),
        "defect_status_counts": defect_status_counts,
        "defect_severity_counts": defect_severity_counts,
        "test_execution_counts": {
            "PASSED": passed,
            "FAILED": failed,
            "BLOCKED": blocked,
            "NOT EXECUTED": not_executed,
        },
        "defect_tracking_rows": defect_tracking_rows,
        "test_execution_rows": test_execution_rows,
    }


def build_report_payload(slug: str) -> dict:
    """Assembles the full report payload for `slug` from files already on disk:

    - streamlit_app/data/<slug>-test-results.json (required -- the run itself)
    - specs/<slug>-test-plan.md (optional -- test steps + objective fallback)
    - streamlit_app/data/<slug>-test-report.json (optional -- narrative text
      authored by Claude Code; falls back to a plain data-derived summary if
      that file doesn't exist yet, so the report is never blank)
    """
    results_path = REPO_ROOT / "streamlit_app" / "data" / f"{slug}-test-results.json"
    results_data = json.loads(results_path.read_text(encoding="utf-8"))

    plan_path = REPO_ROOT / results_data.get("meta", {}).get("test_plan_path", f"specs/{slug}-test-plan.md")
    plan_text = plan_path.read_text(encoding="utf-8") if plan_path.exists() else ""
    plan_steps = parse_test_plan_steps(plan_text) if plan_text else {}
    plan_objective = parse_test_plan_objective(plan_text) if plan_text else ""

    tables = compute_report_tables(results_data, plan_steps)

    narrative_path = REPO_ROOT / "streamlit_app" / "data" / f"{slug}-test-report.json"
    if narrative_path.exists():
        narrative = json.loads(narrative_path.read_text(encoding="utf-8"))
    else:
        summary = results_data.get("summary", {})
        narrative = {
            "test_objectives": plan_objective or "See the test plan for this suite's objectives.",
            "key_findings": (
                f"{summary.get('passed', 0)} of {summary.get('test_cases', 0)} test cases passed "
                f"({summary.get('success_rate', 0)}% success rate). {tables['total_defects']} defect(s) recorded."
            ),
            "recommendations": (
                "Investigate the recorded defects and re-run the suite after fixes land."
                if tables["total_defects"] else
                "No defects recorded -- no immediate action needed."
            ),
            "conclusion": (
                "The suite surfaced real issues that need triage before this feature is considered stable."
                if summary.get("failed", 0) else
                "The suite passed in full; no blocking issues identified in this run."
            ),
        }

    return {
        "slug": slug,
        "title": f"{results_data.get('meta', {}).get('app_under_test', slug)} Test Execution Report",
        "project_name": results_data.get("meta", {}).get("app_under_test", slug),
        "prepared_by": "QA Automation Pipeline (Claude Code)",
        "report_date": results_data.get("meta", {}).get("report_date", ""),
        "testing_period_start": results_data.get("meta", {}).get("report_date", ""),
        "testing_period_end": results_data.get("meta", {}).get("report_date", ""),
        "narrative": narrative,
        **tables,
    }


def build_workbook(payload: dict):
    """Renders `payload` (from build_report_payload) into an openpyxl Workbook
    modeled on this project's adopted Agile Test Report template."""
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    HEADER_FILL = PatternFill("solid", fgColor="9DC3E6")
    SECTION_FILL = PatternFill("solid", fgColor="FFD966")
    BOLD = Font(bold=True)
    WRAP = Alignment(wrap_text=True, vertical="top")

    wb = Workbook()
    ws = wb.active
    ws.title = "Test Execution Report"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 40

    row = 1
    ws.cell(row=row, column=1, value=payload["title"]).font = Font(bold=True, size=16)
    row += 2

    meta_rows = [
        ("Project Name", payload["project_name"]),
        ("Total Defects", payload["total_defects"]),
        ("Total Tests", payload["total_tests"]),
        ("Report Prepared By", payload["prepared_by"]),
        ("Report Date", payload["report_date"]),
        ("Testing Period Start", payload["testing_period_start"]),
        ("Testing Period End", payload["testing_period_end"]),
    ]
    for label, value in meta_rows:
        ws.cell(row=row, column=1, value=label).font = BOLD
        ws.cell(row=row, column=1).fill = HEADER_FILL
        ws.cell(row=row, column=2, value=value)
        row += 1
    row += 1

    for section_title, key in [
        ("Test Objectives", "test_objectives"),
        ("Key Findings", "key_findings"),
        ("Recommendations", "recommendations"),
        ("Conclusion", "conclusion"),
    ]:
        ws.cell(row=row, column=1, value=section_title).font = BOLD
        ws.cell(row=row, column=1).fill = SECTION_FILL
        row += 1
        cell = ws.cell(row=row, column=1, value=payload["narrative"].get(key, ""))
        cell.alignment = WRAP
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.row_dimensions[row].height = 45
        row += 2

    # --- Dashboard Data: hidden-ish side tables the charts bind to ---
    dash_start = row
    ws.cell(row=row, column=1, value="Dashboard Data").font = Font(bold=True, size=13)
    row += 1

    defect_summary_start = row
    ws.cell(row=row, column=1, value="Defect Summary").font = BOLD
    row += 1
    ws.cell(row=row, column=1, value="Status")
    ws.cell(row=row, column=2, value="Count")
    row += 1
    status_rows_start = row
    for status_label, count in payload["defect_status_counts"].items():
        ws.cell(row=row, column=1, value=status_label)
        ws.cell(row=row, column=2, value=count)
        row += 1
    status_rows_end = row - 1
    row += 1

    ws.cell(row=row, column=1, value="Defect Severity").font = BOLD
    row += 1
    ws.cell(row=row, column=1, value="Severity")
    ws.cell(row=row, column=2, value="Count")
    row += 1
    severity_rows_start = row
    for sev_label, count in payload["defect_severity_counts"].items():
        ws.cell(row=row, column=1, value=sev_label)
        ws.cell(row=row, column=2, value=count)
        row += 1
    severity_rows_end = row - 1
    row += 1

    ws.cell(row=row, column=1, value="Test Execution Summary").font = BOLD
    row += 1
    ws.cell(row=row, column=1, value="Status")
    ws.cell(row=row, column=2, value="Count")
    row += 1
    exec_rows_start = row
    for status_label, count in payload["test_execution_counts"].items():
        ws.cell(row=row, column=1, value=status_label)
        ws.cell(row=row, column=2, value=count)
        row += 1
    exec_rows_end = row - 1
    row += 2

    # --- Charts, placed to the right of the dashboard data ---
    density_chart = PieChart()
    density_chart.title = "Defect Density"
    density_data = Reference(ws, min_col=2, min_row=status_rows_start, max_row=status_rows_end)
    density_labels = Reference(ws, min_col=1, min_row=status_rows_start, max_row=status_rows_end)
    density_chart.add_data(density_data, titles_from_data=False)
    density_chart.set_categories(density_labels)
    # Name each slice alongside its percentage, matching the dashboard's pies:
    # a category with no defects would otherwise be an unlabelled 0%.
    density_chart.dataLabels = DataLabelList(showCatName=True, showPercent=True)
    ws.add_chart(density_chart, f"D{defect_summary_start}")

    distribution_chart = PieChart()
    distribution_chart.title = "Defect Distribution"
    dist_data = Reference(ws, min_col=2, min_row=severity_rows_start, max_row=severity_rows_end)
    dist_labels = Reference(ws, min_col=1, min_row=severity_rows_start, max_row=severity_rows_end)
    distribution_chart.add_data(dist_data, titles_from_data=False)
    distribution_chart.set_categories(dist_labels)
    distribution_chart.dataLabels = DataLabelList(showCatName=True, showPercent=True)
    ws.add_chart(distribution_chart, f"D{severity_rows_start}")

    exec_chart = BarChart()
    exec_chart.title = "Test Execution Status"
    exec_chart.type = "col"
    exec_data = Reference(ws, min_col=2, min_row=exec_rows_start, max_row=exec_rows_end)
    exec_labels = Reference(ws, min_col=1, min_row=exec_rows_start, max_row=exec_rows_end)
    exec_chart.add_data(exec_data, titles_from_data=False)
    exec_chart.set_categories(exec_labels)
    ws.add_chart(exec_chart, f"D{exec_rows_start}")

    row = max(row, exec_rows_end + 20)
    row += 2

    # --- Defect Tracking Data ---
    ws.cell(row=row, column=1, value="Defect Tracking Data").font = Font(bold=True, size=13)
    row += 1
    defect_headers = ["Defect ID", "Date Detected", "Description", "Status", "Severity", "Owner", "Remarks"]
    for col, header in enumerate(defect_headers, start=1):
        c = ws.cell(row=row, column=col, value=header)
        c.font = BOLD
        c.fill = HEADER_FILL
    row += 1
    for d in payload["defect_tracking_rows"]:
        ws.cell(row=row, column=1, value=d["id"])
        ws.cell(row=row, column=2, value=d["date_detected"])
        ws.cell(row=row, column=3, value=d["description"])
        ws.cell(row=row, column=4, value=d["status"])
        ws.cell(row=row, column=5, value=d["severity"])
        ws.cell(row=row, column=6, value=d["owner"])
        ws.cell(row=row, column=7, value=d["remarks"])
        row += 1
    row += 2

    # --- Test Execution Data ---
    ws.cell(row=row, column=1, value="Test Execution Data").font = Font(bold=True, size=13)
    row += 1
    test_headers = [
        "Test Case ID", "Date", "Description", "Test Steps", "Expected Result",
        "Actual Result", "Execution Status", "Tested By", "Remarks",
    ]
    for col, header in enumerate(test_headers, start=1):
        c = ws.cell(row=row, column=col, value=header)
        c.font = BOLD
        c.fill = HEADER_FILL
    row += 1
    for t in payload["test_execution_rows"]:
        ws.cell(row=row, column=1, value=t["id"])
        ws.cell(row=row, column=2, value=t["date"])
        ws.cell(row=row, column=3, value=t["description"])
        ws.cell(row=row, column=4, value=t["steps"]).alignment = WRAP
        ws.cell(row=row, column=5, value=t["expected"])
        ws.cell(row=row, column=6, value=t["actual"])
        ws.cell(row=row, column=7, value=t["status"])
        ws.cell(row=row, column=8, value=t["tested_by"])
        ws.cell(row=row, column=9, value=t["remarks"])
        row += 1

    for col in range(3, 10):
        ws.column_dimensions[get_column_letter(col)].width = 24

    return wb


def main() -> None:
    if len(sys.argv) != 2:
        print("Usage: python report_builder.py <slug>", file=sys.stderr)
        sys.exit(1)
    slug = sys.argv[1]

    payload = build_report_payload(slug)
    wb = build_workbook(payload)

    reports_dir = REPO_ROOT / "reports"
    reports_dir.mkdir(exist_ok=True)
    out_path = reports_dir / f"{slug}-test-execution-report.xlsx"
    wb.save(out_path)
    print(f"Wrote {out_path}")


if __name__ == "__main__":
    main()
